#!/usr/bin/env python3
"""Generate data/publications.json from the homepage publication Excel database.

Rules used for the current homepage:
- Use English display fields only.
- Exclude working papers (Type == 工作 / Working / Working Paper).
- Include published and formally accepted records, including Chinese-language records
  when English title/source fields are available.
- DOI links are generated as https://doi.org/<DOI>; records without DOI keep link empty.
- The Language column is preserved so Chinese papers can be excluded from citation display.
"""
from __future__ import annotations
import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("Please install openpyxl first: pip install openpyxl") from exc

TYPE_MAP = {
    "专著": "Monograph",
    "著作": "Monograph",
    "book": "Monograph",
    "monograph": "Monograph",
    "章节": "Book Chapter",
    "book chapter": "Book Chapter",
    "期刊": "Journal Article",
    "journal": "Journal Article",
    "journal article": "Journal Article",
    "会议": "Conference Paper",
    "conference": "Conference Paper",
    "conference paper": "Conference Paper",
}
WORKING_TYPES = {"工作", "working", "working paper", "work", "working papers"}


def clean(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).replace("\r", " ").replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    if s.lower() in {"nan", "none"}:
        return ""
    return s


def norm_key(v: Any) -> str:
    return clean(v).strip().lower()


def norm_name(name: str) -> str:
    s = clean(name).replace("*", "")
    s = re.sub(r"\s+", " ", s)
    if "," in s:
        parts = [p.strip() for p in s.split(",", 1)]
        if len(parts) == 2 and parts[0] and parts[1]:
            s = f"{parts[1]} {parts[0]}"
    return s


def name_key(name: str) -> str:
    return re.sub(r"[^a-z]", "", norm_name(name).lower())


def parse_authors(authors: str, corresponding: str) -> list[str]:
    parts = [p.strip() for p in re.split(r";", clean(authors)) if p.strip()]
    corr_keys = {name_key(p) for p in re.split(r";|, and | and ", clean(corresponding)) if p.strip()}
    # If corresponding_author contains comma-inverted names, splitting by comma can over-split.
    if clean(corresponding):
        corr_keys.add(name_key(corresponding))
        for p in clean(corresponding).split(';'):
            corr_keys.add(name_key(p))
    out = []
    for p in parts:
        n = norm_name(p)
        if name_key(n) in corr_keys and not n.endswith("*"):
            n += "*"
        out.append(n)
    return out


def split_tags(v: Any) -> list[str]:
    s = clean(v)
    if not s:
        return []
    return [x.strip() for x in re.split(r"[,;]", s) if x.strip()]


def doi_link(doi: str) -> tuple[str, str]:
    d = clean(doi)
    if not d:
        return "", ""
    d = re.sub(r"^https?://(dx\.)?doi\.org/", "", d, flags=re.I).strip()
    return d, f"https://doi.org/{d}"


def year_value(v: Any) -> str:
    s = clean(v)
    if not s:
        return ""
    m = re.search(r"\d{4}", s)
    return m.group(0) if m else s


def build_indexes(row: dict[str, Any]) -> list[str]:
    tags = []
    for key in ["SCI", "SSCI", "EI", "CSSCI", "ISTP"]:
        if clean(row.get(key)):
            tags.append(key)
    fms = clean(row.get("FMS"))
    if fms:
        tags.append(f"FMS {fms}" if not fms.upper().startswith("FMS") else fms)
    abs_key = "ABS " if "ABS " in row else "ABS"
    abs_v = clean(row.get(abs_key))
    if abs_v:
        tags.append(f"ABS {abs_v}" if not abs_v.upper().startswith("ABS") else abs_v)
    if clean(row.get("IF")):
        tags.append(f"IF={clean(row.get('IF'))}")
    if clean(row.get("ESI_Highly_Cited")):
        tags.append("ESI Highly Cited Paper")
    if clean(row.get("ESI_Hot")):
        tags.append("ESI Hot Paper")
    # de-duplicate while preserving order
    seen=set(); out=[]
    for t in tags:
        if t not in seen:
            seen.add(t); out.append(t)
    return out


def convert(input_xlsx: Path, output_json: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(input_xlsx, data_only=True)
    ws = wb.active
    headers = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    records = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c-1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        raw_type = norm_key(row.get("Type"))
        if not raw_type or raw_type in WORKING_TYPES:
            continue
        pub_type = TYPE_MAP.get(raw_type, TYPE_MAP.get(clean(row.get("Type")), "Other"))
        title = clean(row.get("Title_English")) or clean(row.get("Title_Chinese"))
        venue = clean(row.get("Source_English")) or clean(row.get("Source_Chinese"))
        authors = parse_authors(clean(row.get("Author_English")) or clean(row.get("Author_Chinese")), clean(row.get("Corresponding_Author")))
        if not (title or venue or authors):
            continue
        doi, link = doi_link(clean(row.get("DOI")))
        language_raw = clean(row.get("Language"))
        language = "zh" if ("中文" in language_raw or language_raw.lower().startswith("zh")) else ("en" if ("英文" in language_raw or language_raw.lower().startswith("en")) else language_raw)
        rec = {
            "type": pub_type,
            "language": language,
            "year": year_value(row.get("Year")),
            "title": title,
            "authors": authors,
            "venue": venue,
            "volume": clean(row.get("Volume")),
            "issue": clean(row.get("Number")),
            "pages": clean(row.get("Page")),
            "doi": doi,
            "link": link,
            "indexes": build_indexes(row),
            "labels": [],
            "note": clean(row.get("Note_English")),
            "isbn": clean(row.get("ISBN")),
            "address": clean(row.get("Address")),
            "conference_date": clean(row.get("Conference_Date")),
            "conference_address": clean(row.get("Conference_Address")),
            "source_row": r,
        }
        records.append(rec)
    type_order = {"Monograph": 0, "Book Chapter": 1, "Journal Article": 2, "Conference Paper": 3, "Other": 9}
    records.sort(key=lambda x: (-(int(x["year"] or 0) if str(x["year"]).isdigit() else 0), type_order.get(x["type"], 9), x.get("venue", ""), x.get("title", "")))
    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    return records


if __name__ == "__main__":
    root = Path(__file__).resolve().parents[1]
    input_xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else root / "publication_database.xlsx"
    output_json = Path(sys.argv[2]) if len(sys.argv) > 2 else root / "data" / "publications.json"
    records = convert(input_xlsx, output_json)
    print(f"Generated {len(records)} publication records -> {output_json}")
