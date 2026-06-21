import os
import re
import sqlite3
import tempfile
import subprocess
import sys
from datetime import datetime

try:
    from pypinyin import lazy_pinyin, Style
except Exception:  # pypinyin is optional but recommended.
    lazy_pinyin = None
    Style = None

import pandas as pd
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "local_content.db")
DEFAULT_EXCEL = os.path.join(DATA_DIR, "homepage_content.xlsx")
DEFAULT_PUBLICATIONS_EXCEL = os.path.join(DATA_DIR, "publication_database.xlsx")
DEFAULT_IMPACT_EXCEL = os.path.join(DATA_DIR, "impact_factors.xlsx")
EXPORT_EXCEL = os.path.join(DATA_DIR, "exported_content.xlsx")
AUTO_EXPORT_HOME_EXCEL = DEFAULT_EXCEL
AUTO_EXPORT_PUBLICATIONS_EXCEL = DEFAULT_PUBLICATIONS_EXCEL
AUTO_EXPORT_EXCEL = os.path.join(DATA_DIR, "auto_exported_content.xlsx")  # legacy combined backup path; not used by the UI
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
SCRIPTS_DIR = os.path.join(BASE_DIR, "scripts")

AUTO_EXPORT_ENABLED = True
SYSTEM_COLUMNS = {"id", "_order_index"}
SKIP_SHEETS = {"README", "Projects", "Project"}
PUBLICATIONS_TABLE = "Publications"
IMPACT_TABLE = "Impact_Factors"
# These tables remain in SQLite and export logic, but are not shown as editable pages in the web UI.
HIDDEN_WEB_TABLES = {IMPACT_TABLE, "Projects", "Project"}
IMPACT_LOOKUP_TABLE = "_impact_lookup"
DIRTY_PUBLICATIONS_TABLE = "_dirty_publications"

app = Flask(__name__)
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)


def connect_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def normalize_identifier(name: str) -> str:
    name = str(name).strip()
    name = re.sub(r"\s+", "_", name)
    name = re.sub(r"[^0-9a-zA-Z_\u4e00-\u9fff]", "_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    if not name:
        name = "Field"
    if re.match(r"^\d", name):
        name = f"C_{name}"
    return name


def table_name_from_sheet(sheet_name: str) -> str:
    return normalize_identifier(sheet_name)


def clean_value(value):
    if pd.isna(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def unique_columns(columns):
    seen = {}
    result = []
    for col in columns:
        base = normalize_identifier(col)
        if base in seen:
            seen[base] += 1
            base = f"{base}_{seen[base]}"
        else:
            seen[base] = 1
        result.append(base)
    return result


def create_meta_table(cur):
    cur.execute(
        'CREATE TABLE IF NOT EXISTS "_meta_tables" ('
        'table_name TEXT PRIMARY KEY, '
        'sheet_name TEXT, '
        'columns TEXT, '
        'source_type TEXT DEFAULT "General", '
        'display_order INTEGER DEFAULT 0)'
    )


def register_meta(cur, table_name, sheet_name, columns, source_type="General", display_order=0):
    cur.execute(
        'INSERT OR REPLACE INTO "_meta_tables" '
        '(table_name, sheet_name, columns, source_type, display_order) VALUES (?, ?, ?, ?, ?)',
        (table_name, sheet_name, "|".join(columns), source_type, display_order),
    )


def drop_user_tables(cur):
    cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
    old_tables = [row[0] for row in cur.fetchall() if not row[0].startswith("sqlite_")]
    for table in old_tables:
        cur.execute(f'DROP TABLE IF EXISTS "{table}"')



def drop_general_tables_preserve_references(cur):
    keep_tables = {PUBLICATIONS_TABLE, IMPACT_TABLE}
    cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
    old_tables = [row[0] for row in cur.fetchall() if not row[0].startswith("sqlite_")]
    for table in old_tables:
        if table not in keep_tables and table != "_meta_tables":
            cur.execute(f'DROP TABLE IF EXISTS "{table}"')
    cur.execute('DELETE FROM "_meta_tables" WHERE table_name NOT IN (?, ?)', (PUBLICATIONS_TABLE, IMPACT_TABLE))


def cleanup_removed_project_tables():
    if not os.path.exists(DB_PATH):
        return
    try:
        conn = connect_db()
        cur = conn.cursor()
        for table in ["Projects", "Project"]:
            cur.execute(f'DROP TABLE IF EXISTS "{table}"')
        cur.execute('DELETE FROM "_meta_tables" WHERE table_name IN (?, ?) OR LOWER(sheet_name) IN (?, ?)',
                    ("Projects", "Project", "projects", "project"))
        conn.commit()
        conn.close()
    except Exception:
        pass


def import_excel_to_tables(excel_path: str, *, mode="append", forced_sheet_name=None, forced_table_name=None, source_type="General", display_order_start=0):
    """Import one workbook into SQLite.

    mode="replace_all": drop all user tables and meta first.
    mode="replace_table": replace only the generated table(s) from this workbook.
    mode="append": add/replace generated table(s) without touching other tables.
    """
    xls = pd.ExcelFile(excel_path)
    conn = connect_db()
    cur = conn.cursor()
    if mode == "replace_all":
        drop_user_tables(cur)
    create_meta_table(cur)

    order = display_order_start
    imported = []
    for sheet_name in xls.sheet_names:
        if sheet_name in SKIP_SHEETS:
            continue
        real_sheet_name = forced_sheet_name if forced_sheet_name else sheet_name
        table_name = forced_table_name if forced_table_name else table_name_from_sheet(real_sheet_name)

        df = pd.read_excel(excel_path, sheet_name=sheet_name, dtype=object)
        df = df.dropna(how="all")
        raw_columns = [c for c in df.columns if str(c).strip() and not str(c).startswith("Unnamed")]
        if not raw_columns:
            continue
        columns = unique_columns(raw_columns)
        df = df.iloc[:, : len(columns)]
        df.columns = columns
        df = df.fillna("")

        cur.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        col_defs = ", ".join([f'"{col}" TEXT' for col in columns])
        cur.execute(f'CREATE TABLE "{table_name}" (id INTEGER PRIMARY KEY AUTOINCREMENT, _order_index INTEGER, {col_defs})')

        placeholders = ", ".join(["?"] * len(columns))
        col_sql = ", ".join([f'"{col}"' for col in columns])
        insert_sql = f'INSERT INTO "{table_name}" (_order_index, {col_sql}) VALUES (?, {placeholders})'

        order_index = 1
        for _, row in df.iterrows():
            values = [clean_value(row[col]) for col in columns]
            if any(v != "" for v in values):
                cur.execute(insert_sql, [order_index] + values)
                order_index += 1

        cur.execute('DELETE FROM "_meta_tables" WHERE table_name = ?', (table_name,))
        register_meta(cur, table_name, real_sheet_name, columns, source_type, order)
        imported.append(table_name)
        order += 1

        if forced_sheet_name or forced_table_name:
            break

    conn.commit()
    conn.close()
    return imported


def ensure_impact_factor_table():
    conn = connect_db()
    cur = conn.cursor()
    create_meta_table(cur)
    row = cur.execute('SELECT table_name FROM "_meta_tables" WHERE table_name = ?', (IMPACT_TABLE,)).fetchone()
    if row is None:
        columns = ["Journal", "Journal_English", "Journal_Chinese", "ISSN", "EISSN", "IF", "JCR_Year", "Quartile", "Category", "Note"]
        col_defs = ", ".join([f'"{col}" TEXT' for col in columns])
        cur.execute(f'CREATE TABLE IF NOT EXISTS "{IMPACT_TABLE}" (id INTEGER PRIMARY KEY AUTOINCREMENT, _order_index INTEGER, {col_defs})')
        max_order = cur.execute('SELECT COALESCE(MAX(display_order), 0) AS m FROM "_meta_tables"').fetchone()["m"]
        register_meta(cur, IMPACT_TABLE, "Impact Factors", columns, "ImpactFactor", int(max_order or 0) + 1)
    conn.commit()
    conn.close()


def database_has_meta():
    if not os.path.exists(DB_PATH) or os.path.getsize(DB_PATH) == 0:
        return False
    try:
        conn = connect_db()
        row = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='_meta_tables'").fetchone()
        conn.close()
        return row is not None
    except sqlite3.DatabaseError:
        return False


def ensure_database():
    if not database_has_meta():
        if os.path.exists(DEFAULT_EXCEL):
            import_excel_to_tables(DEFAULT_EXCEL, mode="replace_all", source_type="Homepage", display_order_start=1)
        else:
            conn = connect_db()
            create_meta_table(conn.cursor())
            conn.commit()
            conn.close()
        if os.path.exists(DEFAULT_PUBLICATIONS_EXCEL):
            # Import publication records as a dedicated table, regardless of original sheet name.
            import_excel_to_tables(
                DEFAULT_PUBLICATIONS_EXCEL,
                mode="append",
                forced_sheet_name="Publications",
                forced_table_name=PUBLICATIONS_TABLE,
                source_type="Publications",
                display_order_start=100,
            )
        if os.path.exists(DEFAULT_IMPACT_EXCEL):
            # Import the local impact-factor database as a dedicated table.
            import_excel_to_tables(
                DEFAULT_IMPACT_EXCEL,
                mode="append",
                forced_sheet_name="Impact Factors",
                forced_table_name=IMPACT_TABLE,
                source_type="ImpactFactor",
                display_order_start=101,
            )
    cleanup_removed_project_tables()
    ensure_impact_factor_table()
    ensure_impact_lookup_table()
    ensure_publication_dirty_table()


def get_tables():
    ensure_database()
    conn = connect_db()
    rows = conn.execute(
        'SELECT table_name, sheet_name, columns, source_type FROM "_meta_tables" '
        'ORDER BY COALESCE(display_order, rowid) ASC, rowid ASC'
    ).fetchall()
    conn.close()
    return [
        {
            "table_name": row["table_name"],
            "sheet_name": row["sheet_name"],
            "columns": row["columns"].split("|") if row["columns"] else [],
            "source_type": row["source_type"] or "General",
        }
        for row in rows
    ]


def get_table_meta(table_name):
    for item in get_tables():
        if item["table_name"] == table_name:
            return item
    return None


def is_hidden_web_table(table_meta):
    table_name = table_meta.get("table_name", "")
    sheet_name = table_meta.get("sheet_name", "")
    if table_name in HIDDEN_WEB_TABLES:
        return True
    if str(sheet_name).strip().lower() in {"project", "projects", "impact factors", "impact_factors"}:
        return True
    return False


def norm_key(value):
    value = str(value or "").strip().lower()
    value = re.sub(r"[\s\-_—–:：,，.;；/\\()（）\[\]{}]+", "", value)
    return value


def normalize_issn(value):
    raw = str(value or "").upper().strip()
    raw = re.sub(r"[^0-9X]", "", raw)
    if len(raw) == 8:
        return raw[:4] + "-" + raw[4:]
    return raw


def col_lookup(columns, candidates):
    norm_map = {}
    for c in columns:
        key = norm_key(c)
        norm_map.setdefault(key, c)
        # normalize_identifier prefixes fields that begin with a digit, e.g.
        # "2025 JIF" becomes "C_2025_JIF". Add a secondary lookup key so the
        # original field name can still be detected.
        if key.startswith("c") and len(key) > 1 and key[1].isdigit():
            norm_map.setdefault(key[1:], c)
    for cand in candidates:
        key = norm_key(cand)
        if key in norm_map:
            return norm_map[key]
    return None


def get_publication_columns(columns):
    """Locate key columns in the Publications table.

    The Publications table intentionally uses a single ISSN field. This field can
    contain either print ISSN or eISSN, so matching treats it as a generic journal
    identifier.
    """
    return {
        "issn": col_lookup(columns, ["ISSN", "Issn", "issn", "EISSN", "eISSN", "E_ISSN", "Print_ISSN", "Online_ISSN"]),
        "if": col_lookup(columns, ["IF", "Impact_Factor", "ImpactFactor", "JIF", "2025_JIF", "2025 JIF"]),
        "source_en": col_lookup(columns, ["Source_English", "Journal_English", "Journal", "Journal_Name", "Journal name", "Source", "刊名"]),
        "source_cn": col_lookup(columns, ["Source_Chinese", "Journal_Chinese", "Chinese_Journal", "期刊名称", "中文刊名"]),
        "title_en": col_lookup(columns, ["Title_English", "Title", "English_Title", "英文题名"]),
        "title_cn": col_lookup(columns, ["Title_Chinese", "Chinese_Title", "中文题名"]),
        "language": col_lookup(columns, ["Language", "Lang", "语言"]),
    }


def get_impact_columns(columns):
    """Locate key columns in the Impact Factors table.

    This supports common JCR field names such as "Journal name", "eISSN",
    "2025 JIF", and "JIF quartile" after Excel columns are normalized.
    """
    return {
        "journal": col_lookup(columns, ["Journal", "Journal_Name", "Journal name", "Source", "Source_English"]),
        "journal_en": col_lookup(columns, ["Journal_English", "English_Name", "Source_English"]),
        "journal_cn": col_lookup(columns, ["Journal_Chinese", "Chinese_Name", "Source_Chinese", "期刊名称", "中文刊名"]),
        "issn": col_lookup(columns, ["ISSN", "Print_ISSN", "P_ISSN", "pISSN"]),
        "eissn": col_lookup(columns, ["EISSN", "eISSN", "E_ISSN", "Online_ISSN", "Electronic_ISSN"]),
        "if": col_lookup(columns, ["IF", "Impact_Factor", "ImpactFactor", "JIF", "Journal_Impact_Factor", "2025 JIF", "2025_JIF"]),
        "jcr_year": col_lookup(columns, ["JCR_Year", "Year"]),
        "quartile": col_lookup(columns, ["Quartile", "JCR_Quartile", "JIF quartile", "JIF_quartile", "Q"]),
    }



def ensure_impact_lookup_table(force_rebuild=False):
    """Create a compact lookup table for fast ISSN/eISSN/journal matching.

    Directly searching the full JCR/impact-factor table on every publication edit
    is slow. This lookup table stores normalized keys and the first matching JCR
    row id, so each single-row update uses indexed exact lookups instead of
    scanning thousands of records.
    """
    if not os.path.exists(DB_PATH):
        return
    conn = connect_db()
    cur = conn.cursor()
    cur.execute(
        f'CREATE TABLE IF NOT EXISTS "{IMPACT_LOOKUP_TABLE}" ('
        'lookup_type TEXT NOT NULL, '
        'lookup_key TEXT NOT NULL, '
        'impact_id INTEGER NOT NULL, '
        'PRIMARY KEY (lookup_type, lookup_key))'
    )
    row_count = cur.execute(f'SELECT COUNT(*) AS c FROM "{IMPACT_LOOKUP_TABLE}"').fetchone()["c"]
    impact_exists = cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (IMPACT_TABLE,)).fetchone()
    if force_rebuild or (impact_exists and int(row_count or 0) == 0):
        rebuild_impact_lookup_table(conn)
    conn.commit()
    conn.close()


def rebuild_impact_lookup_table(conn=None):
    """Rebuild normalized indexes for the hidden impact-factor table.

    For duplicate JCR entries, INSERT OR IGNORE keeps the first record in the
    original Excel/database order, which matches the user's requested behavior.
    """
    own_conn = conn is None
    if own_conn:
        conn = connect_db()
    cur = conn.cursor()
    cur.execute(f'DROP TABLE IF EXISTS "{IMPACT_LOOKUP_TABLE}"')
    cur.execute(
        f'CREATE TABLE "{IMPACT_LOOKUP_TABLE}" ('
        'lookup_type TEXT NOT NULL, '
        'lookup_key TEXT NOT NULL, '
        'impact_id INTEGER NOT NULL, '
        'PRIMARY KEY (lookup_type, lookup_key))'
    )
    meta = conn.execute('SELECT columns FROM "_meta_tables" WHERE table_name = ?', (IMPACT_TABLE,)).fetchone()
    if not meta or not meta["columns"]:
        if own_conn:
            conn.commit(); conn.close()
        return
    cols = get_impact_columns(str(meta["columns"]).split("|"))
    rows = read_table_rows(conn, IMPACT_TABLE)
    insert_sql = f'INSERT OR IGNORE INTO "{IMPACT_LOOKUP_TABLE}" (lookup_type, lookup_key, impact_id) VALUES (?, ?, ?)'
    for row in rows:
        impact_id = row.get("id")
        for c in [cols.get("issn"), cols.get("eissn")]:
            if c:
                key = normalize_issn(row.get(c, ""))
                if key:
                    cur.execute(insert_sql, ("issn", key, impact_id))
        for c in [cols.get("journal"), cols.get("journal_en"), cols.get("journal_cn")]:
            if c:
                key = norm_key(row.get(c, ""))
                if key:
                    cur.execute(insert_sql, ("journal", key, impact_id))
    if own_conn:
        conn.commit()
        conn.close()




def has_table_data_change(table_name):
    return table_name and table_name not in {IMPACT_TABLE, "Projects", "Project", "_meta_tables", IMPACT_LOOKUP_TABLE, DIRTY_PUBLICATIONS_TABLE}


def auto_export_for_table(table_name):
    if not AUTO_EXPORT_ENABLED or not has_table_data_change(table_name):
        return
    if table_name == PUBLICATIONS_TABLE:
        export_publications_excel(AUTO_EXPORT_PUBLICATIONS_EXCEL)
    else:
        export_homepage_excel(AUTO_EXPORT_HOME_EXCEL)

def read_table_rows(conn, table_name):
    return [dict(r) for r in conn.execute(f'SELECT * FROM "{table_name}" ORDER BY COALESCE(_order_index, id) ASC, id ASC').fetchall()]


def get_impact_record_by_lookup(conn, lookup_type, lookup_key):
    if not lookup_key:
        return None
    lookup = conn.execute(
        f'SELECT impact_id FROM "{IMPACT_LOOKUP_TABLE}" WHERE lookup_type = ? AND lookup_key = ? LIMIT 1',
        (lookup_type, lookup_key),
    ).fetchone()
    if not lookup:
        return None
    row = conn.execute(f'SELECT * FROM "{IMPACT_TABLE}" WHERE id = ? LIMIT 1', (lookup["impact_id"],)).fetchone()
    return dict(row) if row else None


def find_impact_record(conn, pub_row, pub_cols):
    """Find one impact-factor record for a publication row using the lookup table.

    Matching order:
    1. Exact normalized journal name from Source_English / Source_Chinese.
    2. Exact normalized ISSN/eISSN from the single Publications ISSN field.

    This intentionally avoids fuzzy matching because fuzzy scans are slow and can
    incorrectly backfill IF values for similarly named journals.
    """
    for c in [pub_cols.get("source_en"), pub_cols.get("source_cn")]:
        if c:
            key = norm_key(pub_row.get(c, ""))
            impact_row = get_impact_record_by_lookup(conn, "journal", key)
            if impact_row:
                return impact_row, "Journal"

    current_identifier = pub_row.get(pub_cols.get("issn") or "", "")
    key = normalize_issn(current_identifier)
    impact_row = get_impact_record_by_lookup(conn, "issn", key)
    if impact_row:
        return impact_row, "ISSN/EISSN"
    return None, ""

def update_single_publication_match(conn, pub_id, *, overwrite_if=False, ensure_lookup=True):
    """Auto-fill ISSN/EISSN and IF for one publication record only.

    The function is called after a single Publications row is added or edited.
    It does not scan or rewrite the whole Publications table.
    """
    pub_meta = get_table_meta(PUBLICATIONS_TABLE)
    impact_meta = get_table_meta(IMPACT_TABLE)
    if not pub_meta or not impact_meta:
        return {"matched": False, "reason": "Missing Publications or Impact Factors table"}

    pub_cols = get_publication_columns(pub_meta["columns"])
    impact_cols = get_impact_columns(impact_meta["columns"])
    if not pub_cols.get("issn"):
        return {"matched": False, "reason": "Publications table must contain an ISSN column"}
    if not pub_cols.get("if"):
        return {"matched": False, "reason": "Publications table must contain an IF column"}
    if not impact_cols.get("if"):
        return {"matched": False, "reason": "Impact Factors table must contain an IF/JIF column"}

    if ensure_lookup:
        ensure_impact_lookup_table()
    pub_row = conn.execute(f'SELECT * FROM "{PUBLICATIONS_TABLE}" WHERE id = ?', (pub_id,)).fetchone()
    if not pub_row:
        return {"matched": False, "reason": "Publication row not found"}
    pub_row = dict(pub_row)
    impact_row, method = find_impact_record(conn, pub_row, pub_cols)
    if not impact_row:
        return {"matched": False, "reason": "No matching journal name, ISSN, or eISSN in Impact Factors"}

    updates = {}

    current_pub_identifier = str(pub_row.get(pub_cols["issn"], "") or "").strip()
    if overwrite_if or not current_pub_identifier:
        # Prefer print ISSN; if it is empty, use eISSN. Both are stored in the single Publications ISSN field.
        for c in [impact_cols.get("issn"), impact_cols.get("eissn")]:
            value = str(impact_row.get(c, "") or "").strip() if c else ""
            if value:
                new_identifier = normalize_issn(value)
                if new_identifier != normalize_issn(current_pub_identifier):
                    updates[pub_cols["issn"]] = new_identifier
                break

    impact_if = str(impact_row.get(impact_cols["if"], "") or "").strip()
    current_if = str(pub_row.get(pub_cols["if"], "") or "").strip()
    if impact_if and (overwrite_if or not current_if):
        updates[pub_cols["if"]] = impact_if

    if updates:
        set_sql = ", ".join([f'"{c}" = ?' for c in updates])
        conn.execute(f'UPDATE "{PUBLICATIONS_TABLE}" SET {set_sql} WHERE id = ?', list(updates.values()) + [pub_id])
    return {"matched": True, "method": method, "updates": updates}

def auto_match_publication_by_id(pub_id, *, overwrite_if=False):
    try:
        conn = connect_db()
        result = update_single_publication_match(conn, pub_id, overwrite_if=overwrite_if)
        conn.commit()
        conn.close()
        return result
    except Exception as exc:
        return {"matched": False, "reason": str(exc)}


def ensure_publication_dirty_table(conn=None):
    """Track Publications rows that were newly added or edited and need one-click enrichment."""
    own_conn = conn is None
    if own_conn:
        conn = connect_db()
    conn.execute(
        f'CREATE TABLE IF NOT EXISTS "{DIRTY_PUBLICATIONS_TABLE}" ('
        'pub_id INTEGER PRIMARY KEY, '
        'updated_at TEXT DEFAULT CURRENT_TIMESTAMP)'
    )
    if own_conn:
        conn.commit()
        conn.close()


def mark_publication_dirty(conn, pub_id):
    if not pub_id:
        return
    ensure_publication_dirty_table(conn)
    conn.execute(
        f'INSERT OR REPLACE INTO "{DIRTY_PUBLICATIONS_TABLE}" (pub_id, updated_at) VALUES (?, CURRENT_TIMESTAMP)',
        (pub_id,),
    )


def clear_publication_dirty(conn, pub_ids=None):
    ensure_publication_dirty_table(conn)
    if pub_ids is None:
        conn.execute(f'DELETE FROM "{DIRTY_PUBLICATIONS_TABLE}"')
    else:
        ids = [int(x) for x in pub_ids if x]
        if ids:
            placeholders = ','.join(['?'] * len(ids))
            conn.execute(f'DELETE FROM "{DIRTY_PUBLICATIONS_TABLE}" WHERE pub_id IN ({placeholders})', ids)


def get_dirty_publication_ids(conn):
    ensure_publication_dirty_table(conn)
    # Keep only rows that still exist in Publications.
    if not get_table_meta(PUBLICATIONS_TABLE):
        return []
    conn.execute(
        f'DELETE FROM "{DIRTY_PUBLICATIONS_TABLE}" '
        f'WHERE pub_id NOT IN (SELECT id FROM "{PUBLICATIONS_TABLE}")'
    )
    rows = conn.execute(
        f'SELECT pub_id FROM "{DIRTY_PUBLICATIONS_TABLE}" ORDER BY updated_at ASC, pub_id ASC'
    ).fetchall()
    return [int(r["pub_id"]) for r in rows]


def mark_all_publications_dirty(conn):
    ensure_publication_dirty_table(conn)
    if not get_table_meta(PUBLICATIONS_TABLE):
        return 0
    rows = conn.execute(f'SELECT id FROM "{PUBLICATIONS_TABLE}"').fetchall()
    for row in rows:
        mark_publication_dirty(conn, int(row["id"]))
    return len(rows)


COMPOUND_SURNAMES = [
    "欧阳", "司马", "上官", "诸葛", "东方", "独孤", "南宫", "万俟", "闻人", "夏侯",
    "尉迟", "公羊", "赫连", "澹台", "皇甫", "宗政", "濮阳", "公冶", "太叔", "申屠",
    "公孙", "慕容", "仲孙", "钟离", "长孙", "宇文", "司徒", "鲜于", "司空", "闾丘",
    "子车", "亓官", "司寇", "巫马", "公西", "颛孙", "壤驷", "公良", "漆雕", "乐正",
    "宰父", "谷梁", "拓跋", "夹谷", "轩辕", "令狐", "段干", "百里", "呼延", "东郭",
    "南门", "羊舌", "微生", "公户", "公玉", "公仪", "梁丘", "公仲", "公上", "公门",
]


def split_author_names(text):
    text = str(text or "").strip()
    if not text:
        return []
    text = re.sub(r"[；;|/]+", "、", text)
    text = re.sub(r"\s*,\s*", "、", text)
    text = re.sub(r"\s+", "", text)
    return [x.strip() for x in re.split(r"[、，]+", text) if x.strip()]


def chinese_name_to_english(name):
    name = str(name or "").strip()
    name = re.sub(r"\s+", "", name)
    if not name:
        return ""
    # If it already contains Latin characters, preserve it.
    if re.search(r"[A-Za-z]", name):
        return name
    if lazy_pinyin is None:
        return ""
    surname = name[:1]
    given = name[1:]
    for cs in COMPOUND_SURNAMES:
        if name.startswith(cs):
            surname = cs
            given = name[len(cs):]
            break
    surname_py = " ".join(lazy_pinyin(surname)).title()
    given_py = "".join(lazy_pinyin(given)).title() if given else ""
    if given_py:
        return f"{surname_py}, {given_py}"
    return surname_py


def chinese_authors_to_english(text):
    names = split_author_names(text)
    converted = [chinese_name_to_english(n) for n in names]
    converted = [x for x in converted if x]
    return "; ".join(converted)


def auto_fill_publication_authors(conn, pub_id):
    """Fill Author_English and Corresponding_Author_en for one row when empty.

    This is a local pinyin transliteration, not machine translation. It keeps
    manually entered English names unchanged.
    """
    meta = get_table_meta(PUBLICATIONS_TABLE)
    if not meta:
        return {"updated": False, "reason": "Missing Publications table"}
    cols = meta["columns"]
    cn_author = col_lookup(cols, ["Author_Chinese", "Authors_Chinese", "Chinese_Author", "作者", "中文作者"])
    en_author = col_lookup(cols, ["Author_English", "Authors_English", "English_Author", "英文作者"])
    cn_corr = col_lookup(cols, ["Corresponding_Author_cn", "Corresponding_Author_CN", "Chinese_Corresponding_Author", "通讯作者中文"])
    en_corr = col_lookup(cols, ["Corresponding_Author_en", "Corresponding_Author_EN", "English_Corresponding_Author", "通讯作者英文"])
    if not ((cn_author and en_author) or (cn_corr and en_corr)):
        return {"updated": False, "reason": "Missing author columns"}
    row = conn.execute(f'SELECT * FROM "{PUBLICATIONS_TABLE}" WHERE id = ?', (pub_id,)).fetchone()
    if not row:
        return {"updated": False, "reason": "Publication row not found"}
    row = dict(row)
    updates = {}
    if cn_author and en_author and not str(row.get(en_author, "") or "").strip():
        converted = chinese_authors_to_english(row.get(cn_author, ""))
        if converted:
            updates[en_author] = converted
    if cn_corr and en_corr and not str(row.get(en_corr, "") or "").strip():
        converted = chinese_authors_to_english(row.get(cn_corr, ""))
        if converted:
            updates[en_corr] = converted
    if updates:
        set_sql = ", ".join([f'"{c}" = ?' for c in updates])
        conn.execute(f'UPDATE "{PUBLICATIONS_TABLE}" SET {set_sql} WHERE id = ?', list(updates.values()) + [pub_id])
        return {"updated": True, "updates": updates}
    return {"updated": False, "reason": "English author fields already filled or no Chinese names"}




def is_english_publication_language(value):
    value = str(value or "").strip().lower()
    if not value:
        return False
    return value in {"英文", "english", "en", "eng"}


def auto_fill_publication_title(conn, pub_id):
    """For English publications, copy Title_English into Title_Chinese when the Chinese title is empty.

    The Chinese title field is sometimes required by downstream homepage/Excel templates.
    For English papers, the requested default is to reuse the English title rather than
    invoking machine translation. Existing Chinese titles are preserved.
    """
    meta = get_table_meta(PUBLICATIONS_TABLE)
    if not meta:
        return {"updated": False, "reason": "Missing Publications table"}
    cols = get_publication_columns(meta["columns"])
    title_en = cols.get("title_en")
    title_cn = cols.get("title_cn")
    language = cols.get("language")
    if not (title_en and title_cn):
        return {"updated": False, "reason": "Missing title columns"}
    row = conn.execute(f'SELECT * FROM "{PUBLICATIONS_TABLE}" WHERE id = ?', (pub_id,)).fetchone()
    if not row:
        return {"updated": False, "reason": "Publication row not found"}
    row = dict(row)
    en_title = str(row.get(title_en, "") or "").strip()
    cn_title = str(row.get(title_cn, "") or "").strip()
    lang_value = str(row.get(language, "") or "").strip() if language else ""
    if en_title and not cn_title and (not language or is_english_publication_language(lang_value)):
        conn.execute(f'UPDATE "{PUBLICATIONS_TABLE}" SET "{title_cn}" = ? WHERE id = ?', (en_title, pub_id))
        return {"updated": True, "updates": {title_cn: en_title}}
    return {"updated": False, "reason": "Chinese title already filled, English title missing, or language is not English"}


def auto_fill_publication_source(conn, pub_id):
    """For English publications, copy Source_English into Source_Chinese when empty.

    For English papers, the Chinese source field can reuse the English journal/source
    name so downstream Excel/homepage templates always have a display value. Existing
    manually entered Chinese source names are preserved.
    """
    meta = get_table_meta(PUBLICATIONS_TABLE)
    if not meta:
        return {"updated": False, "reason": "Missing Publications table"}
    cols = get_publication_columns(meta["columns"])
    source_en = cols.get("source_en")
    source_cn = cols.get("source_cn")
    language = cols.get("language")
    if not (source_en and source_cn):
        return {"updated": False, "reason": "Missing source columns"}
    row = conn.execute(f'SELECT * FROM "{PUBLICATIONS_TABLE}" WHERE id = ?', (pub_id,)).fetchone()
    if not row:
        return {"updated": False, "reason": "Publication row not found"}
    row = dict(row)
    en_source = str(row.get(source_en, "") or "").strip()
    cn_source = str(row.get(source_cn, "") or "").strip()
    lang_value = str(row.get(language, "") or "").strip() if language else ""
    if en_source and not cn_source and (not language or is_english_publication_language(lang_value)):
        conn.execute(f'UPDATE "{PUBLICATIONS_TABLE}" SET "{source_cn}" = ? WHERE id = ?', (en_source, pub_id))
        return {"updated": True, "updates": {source_cn: en_source}}
    return {"updated": False, "reason": "Chinese source already filled, English source missing, or language is not English"}


def style_excel_file(path):
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="3F745F")
    header_font = Font(color="FFFFFF", bold=True)
    soft_fill = PatternFill("solid", fgColor="FCFBF7")
    highlight_fill = PatternFill("solid", fgColor="FFF2E6")
    thin = Side(style="thin", color="E4DFD3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row == 0 or max_col == 0:
            continue
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.row % 2 == 0:
                    cell.fill = soft_fill
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            header = str(ws.cell(row=1, column=col_idx).value or "")
            values = [str(ws.cell(row=r, column=col_idx).value or "") for r in range(1, min(max_row, 60) + 1)]
            max_len = max([len(v) for v in values] + [len(header)])
            if any(k in header.lower() for k in ["title", "author", "abstract", "note", "description", "link", "doi"]):
                width = min(max(max_len + 2, 18), 48)
            else:
                width = min(max(max_len + 2, 10), 24)
            ws.column_dimensions[letter].width = width
            if header.upper() in {"IF", "DOI", "ISSN", "EISSN"}:
                # Make key identifier columns a little warmer without repeatedly
                # reassigning existing StyleProxy objects.
                ws.cell(row=1, column=col_idx).fill = highlight_fill
                ws.cell(row=1, column=col_idx).font = Font(color="5A3E1B", bold=True)
        ws.row_dimensions[1].height = 28
    wb.save(path)


def export_tables_to_excel(output_path: str, tables):
    conn = connect_db()
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        wrote_any = False
        for table in tables:
            df = pd.read_sql_query(
                f'SELECT * FROM "{table["table_name"]}" ORDER BY COALESCE(_order_index, id) ASC, id ASC', conn
            )
            for system_col in ["id", "_order_index"]:
                if system_col in df.columns:
                    df = df.drop(columns=[system_col])
            df.to_excel(writer, sheet_name=table["sheet_name"][:31], index=False)
            wrote_any = True
        if not wrote_any:
            pd.DataFrame().to_excel(writer, sheet_name="Empty", index=False)
    conn.close()
    style_excel_file(output_path)
    return output_path


def homepage_export_tables():
    return [
        t for t in get_tables()
        if t["table_name"] not in {PUBLICATIONS_TABLE, IMPACT_TABLE, "Projects", "Project"}
        and str(t.get("sheet_name", "")).strip().lower() not in {"projects", "project", "impact factors", "impact_factors"}
    ]


def publication_export_tables():
    return [t for t in get_tables() if t["table_name"] == PUBLICATIONS_TABLE]


def export_homepage_excel(output_path: str = AUTO_EXPORT_HOME_EXCEL):
    return export_tables_to_excel(output_path, homepage_export_tables())


def export_publications_excel(output_path: str = AUTO_EXPORT_PUBLICATIONS_EXCEL):
    return export_tables_to_excel(output_path, publication_export_tables())


def export_database_to_excel(output_path: str, *, include_impact_factors=False):
    # Legacy combined backup export. The UI now downloads the two source workbooks separately.
    tables = homepage_export_tables() + publication_export_tables()
    if include_impact_factors:
        tables += [t for t in get_tables() if t["table_name"] == IMPACT_TABLE]
    return export_tables_to_excel(output_path, tables)


def auto_export_excel():
    # Full sync is used only on startup or explicit backup operations.
    if AUTO_EXPORT_ENABLED:
        export_homepage_excel(AUTO_EXPORT_HOME_EXCEL)
        export_publications_excel(AUTO_EXPORT_PUBLICATIONS_EXCEL)


def create_or_update_row(table_name, row_id=None, payload=None):
    meta = get_table_meta(table_name)
    if not meta:
        return None, None, ({"error": "Table not found"}, 404)
    payload = payload or {}
    cols = meta["columns"]
    values = [str(payload.get(col, "")).strip() for col in cols]
    conn = connect_db()
    cur = conn.cursor()
    if row_id is None:
        min_order = cur.execute(f'SELECT COALESCE(MIN(_order_index), 0) AS min_order FROM "{table_name}"').fetchone()["min_order"]
        # New records are inserted at the top of every table. Existing Excel order is preserved
        # because display/export sorting is ascending by _order_index, id.
        new_order = int(min_order or 0) - 1
        placeholders = ", ".join(["?"] * len(cols))
        col_sql = ", ".join([f'"{c}"' for c in cols])
        insert_sql = f'INSERT INTO "{table_name}" (_order_index, {col_sql}) VALUES (?, {placeholders})'
        cur.execute(insert_sql, [new_order] + values)
        row_id = cur.lastrowid
    else:
        set_sql = ", ".join([f'"{c}" = ?' for c in cols])
        cur.execute(f'UPDATE "{table_name}" SET {set_sql} WHERE id = ?', values + [row_id])

    if table_name == PUBLICATIONS_TABLE:
        mark_publication_dirty(conn, row_id)
    conn.commit()
    conn.close()
    auto_export_for_table(table_name)
    return row_id, None, None


@app.route("/")
def index():
    ensure_database()
    return render_template("index.html")


@app.route("/api/tables")
def api_tables():
    # Hide auxiliary/reference tables from the sidebar while keeping them in SQLite.
    return jsonify([t for t in get_tables() if not is_hidden_web_table(t)])


@app.route("/api/rows/<table_name>")
def api_rows(table_name):
    meta = get_table_meta(table_name)
    if not meta:
        return jsonify({"error": "Table not found"}), 404
    q = request.args.get("q", "").strip()
    conn = connect_db()
    if q:
        conditions = " OR ".join([f'"{c}" LIKE ?' for c in meta["columns"]])
        params = [f"%{q}%"] * len(meta["columns"])
        rows = conn.execute(
            f'SELECT * FROM "{table_name}" WHERE {conditions} ORDER BY COALESCE(_order_index, id) ASC, id ASC', params
        ).fetchall()
    else:
        rows = conn.execute(f'SELECT * FROM "{table_name}" ORDER BY COALESCE(_order_index, id) ASC, id ASC').fetchall()
    conn.close()
    return jsonify([dict(row) for row in rows])


@app.route("/api/options/<table_name>/<column_name>")
def api_options(table_name, column_name):
    meta = get_table_meta(table_name)
    if not meta or column_name not in meta["columns"]:
        return jsonify([])
    conn = connect_db()
    rows = conn.execute(
        f'SELECT DISTINCT "{column_name}" AS value FROM "{table_name}" WHERE TRIM(COALESCE("{column_name}", "")) <> "" ORDER BY value LIMIT 80'
    ).fetchall()
    conn.close()
    values = [row["value"] for row in rows if str(row["value"]).strip()]
    # Add common yes/no choices for index fields even before values exist.
    if norm_key(column_name) in {norm_key(x) for x in ["SCI", "SSCI", "EI", "CSSCI", "ISTP", "ESI_Highly_Cited", "ESI_Hot"]}:
        for v in ["是", "否"]:
            if v not in values:
                values.append(v)
    if norm_key(column_name) == norm_key("Language"):
        for v in ["中文", "英文"]:
            if v not in values:
                values.append(v)
    if norm_key(column_name) == norm_key("Type"):
        for v in ["期刊", "会议", "专著", "章节", "工作"]:
            if v not in values:
                values.append(v)
    return jsonify(values)


@app.route("/api/rows/<table_name>", methods=["POST"])
def api_add_row(table_name):
    row_id, match_result, err = create_or_update_row(table_name, None, request.get_json(force=True))
    if err:
        payload, status = err
        return jsonify(payload), status
    return jsonify({"ok": True, "id": row_id, "match_result": match_result})


@app.route("/api/rows/<table_name>/<int:row_id>", methods=["PUT"])
def api_update_row(table_name, row_id):
    _, match_result, err = create_or_update_row(table_name, row_id, request.get_json(force=True))
    if err:
        payload, status = err
        return jsonify(payload), status
    return jsonify({"ok": True, "match_result": match_result})


@app.route("/api/rows/<table_name>/<int:row_id>", methods=["DELETE"])
def api_delete_row(table_name, row_id):
    meta = get_table_meta(table_name)
    if not meta:
        return jsonify({"error": "Table not found"}), 404
    conn = connect_db()
    conn.execute(f'DELETE FROM "{table_name}" WHERE id = ?', [row_id])
    if table_name == PUBLICATIONS_TABLE:
        clear_publication_dirty(conn, [row_id])
    conn.commit()
    conn.close()
    auto_export_for_table(table_name)
    return jsonify({"ok": True})


@app.route("/api/import", methods=["POST"])
def api_import_excel():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Please upload an Excel file"}), 400
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name
    try:
        conn = connect_db()
        cur = conn.cursor()
        create_meta_table(cur)
        drop_general_tables_preserve_references(cur)
        conn.commit()
        conn.close()
        import_excel_to_tables(tmp_path, mode="append", source_type="Homepage", display_order_start=1)
        ensure_impact_factor_table()
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    export_homepage_excel(AUTO_EXPORT_HOME_EXCEL)
    return jsonify({"ok": True})


@app.route("/api/import_publications", methods=["POST"])
def api_import_publications():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name
    try:
        import_excel_to_tables(
            tmp_path,
            mode="append",
            forced_sheet_name="Publications",
            forced_table_name=PUBLICATIONS_TABLE,
            source_type="Publications",
            display_order_start=100,
        )
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    conn = connect_db()
    mark_all_publications_dirty(conn)
    conn.commit()
    conn.close()
    export_publications_excel(AUTO_EXPORT_PUBLICATIONS_EXCEL)
    return jsonify({"ok": True})


@app.route("/api/import_impact_factors", methods=["POST"])
def api_import_impact_factors():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name
    try:
        import_excel_to_tables(
            tmp_path,
            mode="append",
            forced_sheet_name="Impact Factors",
            forced_table_name=IMPACT_TABLE,
            source_type="ImpactFactor",
            display_order_start=101,
        )
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    conn = connect_db()
    rebuild_impact_lookup_table(conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/publications/match_if/<int:pub_id>", methods=["POST"])
def api_match_publication_if(pub_id):
    """Manually re-match one publication row."""
    ensure_database()
    overwrite = str(request.args.get("overwrite", "0")).lower() in {"1", "true", "yes"}
    conn = connect_db()
    result = update_single_publication_match(conn, pub_id, overwrite_if=overwrite)
    conn.commit()
    conn.close()
    export_publications_excel(AUTO_EXPORT_PUBLICATIONS_EXCEL)
    return jsonify({"ok": True, "result": result})


@app.route("/api/publications/translate_authors/<int:pub_id>", methods=["POST"])
def api_translate_publication_authors(pub_id):
    """Generate Author_English / Corresponding_Author_en from Chinese names for one row."""
    ensure_database()
    conn = connect_db()
    result = auto_fill_publication_authors(conn, pub_id)
    conn.commit()
    conn.close()
    export_publications_excel(AUTO_EXPORT_PUBLICATIONS_EXCEL)
    return jsonify({"ok": True, "result": result})


@app.route("/api/publications/dirty_count")
def api_publications_dirty_count():
    ensure_database()
    conn = connect_db()
    ids = get_dirty_publication_ids(conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "count": len(ids)})


@app.route("/api/publications/update_changed", methods=["POST"])
def api_update_changed_publications():
    """Batch update only Publications rows that were added or edited since last update."""
    ensure_database()
    ensure_impact_lookup_table()
    conn = connect_db()
    dirty_ids = get_dirty_publication_ids(conn)
    results = []
    changed = False
    for pub_id in dirty_ids:
        author_result = auto_fill_publication_authors(conn, pub_id)
        title_result = auto_fill_publication_title(conn, pub_id)
        source_result = auto_fill_publication_source(conn, pub_id)
        impact_result = update_single_publication_match(conn, pub_id, overwrite_if=True, ensure_lookup=False)
        row_changed = (
            bool(author_result.get("updated"))
            or bool(title_result.get("updated"))
            or bool(source_result.get("updated"))
            or bool(impact_result.get("updates"))
        )
        changed = changed or row_changed
        results.append({
            "id": pub_id,
            "authors": author_result,
            "title": title_result,
            "source": source_result,
            "impact": impact_result,
            "changed": row_changed,
        })
    clear_publication_dirty(conn, dirty_ids)
    conn.commit()
    conn.close()
    if changed:
        export_publications_excel(AUTO_EXPORT_PUBLICATIONS_EXCEL)
    matched = sum(1 for item in results if item.get("impact", {}).get("matched"))
    return jsonify({
        "ok": True,
        "processed": len(dirty_ids),
        "changed": changed,
        "matched": matched,
        "results": results,
    })


@app.route("/api/publications/update_fields/<int:pub_id>", methods=["POST"])
def api_update_publication_fields(pub_id):
    """One-click update for a single publication row.

    It fills English author names when empty and refreshes ISSN/IF from the hidden
    impact-factor table. ISSN/IF are overwritten when a match is found, so editing
    a journal name and clicking 更新 immediately refreshes the current row.
    """
    ensure_database()
    conn = connect_db()
    author_result = auto_fill_publication_authors(conn, pub_id)
    title_result = auto_fill_publication_title(conn, pub_id)
    source_result = auto_fill_publication_source(conn, pub_id)
    impact_result = update_single_publication_match(conn, pub_id, overwrite_if=True)
    conn.commit()
    conn.close()
    changed = (
        bool(author_result.get("updated"))
        or bool(title_result.get("updated"))
        or bool(source_result.get("updated"))
        or bool(impact_result.get("updates"))
    )
    if changed:
        export_publications_excel(AUTO_EXPORT_PUBLICATIONS_EXCEL)
    return jsonify({"ok": True, "result": {"authors": author_result, "title": title_result, "source": source_result, "impact": impact_result, "changed": changed}})


@app.route("/api/export")
def api_export_excel():
    # Legacy combined workbook download for backup.
    export_database_to_excel(EXPORT_EXCEL, include_impact_factors=False)
    return send_file(EXPORT_EXCEL, as_attachment=True, download_name="exported_content.xlsx")


@app.route("/api/export/homepage")
def api_export_homepage_excel():
    export_homepage_excel(AUTO_EXPORT_HOME_EXCEL)
    return send_file(AUTO_EXPORT_HOME_EXCEL, as_attachment=True, download_name="homepage_content.xlsx")


@app.route("/api/export/publications")
def api_export_publications_excel():
    export_publications_excel(AUTO_EXPORT_PUBLICATIONS_EXCEL)
    return send_file(AUTO_EXPORT_PUBLICATIONS_EXCEL, as_attachment=True, download_name="publication_database.xlsx")


@app.route("/api/auto_export_status")
def api_auto_export_status():
    paths = [AUTO_EXPORT_HOME_EXCEL, AUTO_EXPORT_PUBLICATIONS_EXCEL]
    existing = [p for p in paths if os.path.exists(p)]
    modified_at = ""
    if existing:
        modified_at = datetime.fromtimestamp(max(os.path.getmtime(p) for p in existing)).strftime("%Y-%m-%d %H:%M:%S")
    return jsonify({"enabled": AUTO_EXPORT_ENABLED, "exists": bool(existing), "paths": paths, "modified_at": modified_at})


def run_local_script(script_name: str):
    allowed = {
        "json": "generate_json.py",
        "tex": "generate_tex.py",
        "json_tex": "generate_all.py",
    }
    if script_name not in allowed:
        return {"ok": False, "error": "Unknown task"}, 400
    script_path = os.path.join(SCRIPTS_DIR, allowed[script_name])
    if not os.path.exists(script_path):
        return {"ok": False, "error": f"Script not found: {allowed[script_name]}"}, 404
    try:
        result = subprocess.run(
            [sys.executable, script_path, "--db", DB_PATH, "--out", OUTPUT_DIR],
            cwd=BASE_DIR,
            text=True,
            capture_output=True,
            timeout=60,
            check=False,
        )
        return {
            "ok": result.returncode == 0,
            "task": script_name,
            "returncode": result.returncode,
            "stdout": result.stdout.strip(),
            "stderr": result.stderr.strip(),
        }, 200 if result.returncode == 0 else 500
    except subprocess.TimeoutExpired:
        return {"ok": False, "error": "Script timed out after 60 seconds"}, 500


@app.route("/api/run/<task>", methods=["POST"])
def api_run_task(task):
    payload, status = run_local_script(task)
    return jsonify(payload), status


@app.route("/api/files")
def api_files():
    files = []
    for name in sorted(os.listdir(OUTPUT_DIR)):
        path = os.path.join(OUTPUT_DIR, name)
        if os.path.isfile(path):
            files.append({
                "name": name,
                "size": os.path.getsize(path),
                "modified_at": datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M:%S"),
                "url": f"/api/files/{name}",
            })
    return jsonify(files)


@app.route("/api/files/<path:filename>")
def api_download_generated_file(filename):
    safe_name = os.path.basename(filename)
    path = os.path.join(OUTPUT_DIR, safe_name)
    if not os.path.exists(path):
        return jsonify({"error": "File not found"}), 404
    return send_file(path, as_attachment=True, download_name=safe_name)


if __name__ == "__main__":
    ensure_database()
    app.run(host="127.0.0.1", port=5000, debug=True)
