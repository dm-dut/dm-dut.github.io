from __future__ import annotations

import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook

from .config import JOURNAL_LIST_PATH, WHITELIST_REQUIRED
from .utils import normalize_space


_PROVIDER_ALIASES = {
    "elsevier": "sciencedirect",
    "sciencedirect": "sciencedirect",
    "science direct": "sciencedirect",
    "springer": "springer",
    "springer nature": "springer",
    "springernature": "springer",
    "ieee": "ieee",
    "ieee xplore": "ieee",
}

_PUBLISHER_LABELS = {
    "sciencedirect": "Elsevier",
    "springer": "Springer Nature",
    "ieee": "IEEE",
}


@dataclass(frozen=True)
class JournalSpec:
    provider: str
    publisher: str
    journal: str
    issns: tuple[str, ...] = ()
    aliases: tuple[str, ...] = ()
    category: str = ""
    notes: str = ""


def normalize_issn(value: str | None) -> str:
    """Normalize ISSN/eISSN to eight alphanumeric characters without a hyphen."""
    return re.sub(r"[^0-9Xx]", "", str(value or "")).upper()


def display_issn(value: str) -> str:
    value = normalize_issn(value)
    return f"{value[:4]}-{value[4:]}" if len(value) == 8 else value


def _truthy(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on", "enabled", "是"}


def _provider(value: str | None) -> str:
    key = normalize_space(value).lower()
    return _PROVIDER_ALIASES.get(key, key)


def _title_key(value: str | None) -> str:
    text = normalize_space(value).lower()
    # Ignore punctuation that often varies between API metadata and journal lists.
    return re.sub(r"[^a-z0-9]+", "", text)


def _split_aliases(value: str | None) -> tuple[str, ...]:
    if not value:
        return ()
    return tuple(x for x in (normalize_space(v) for v in re.split(r"[;|\n]", str(value))) if x)


def _header_key(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


@lru_cache(maxsize=1)
def load_journal_list(path: str | Path | None = None) -> tuple[JournalSpec, ...]:
    target = Path(path or JOURNAL_LIST_PATH)
    if not target.exists():
        if WHITELIST_REQUIRED:
            raise FileNotFoundError(
                f"Journal whitelist not found: {target}. Create journal_list.xlsx or set JOURNAL_LIST_PATH."
            )
        return ()

    wb = load_workbook(target, read_only=True, data_only=True)
    ws = wb["Journals"] if "Journals" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return ()

    col = {_header_key(name): idx for idx, name in enumerate(header)}

    def val(row, *names):
        for name in names:
            idx = col.get(_header_key(name))
            if idx is not None and idx < len(row):
                v = row[idx]
                if v not in (None, ""):
                    return v
        return ""

    specs: list[JournalSpec] = []
    for row in rows:
        if not _truthy(val(row, "Enabled", "Enable", "Active")):
            continue

        provider = _provider(str(val(row, "Provider", "Publisher", "Source")))
        journal = normalize_space(str(val(row, "Journal", "Journal Title", "Title")))
        if provider not in _PUBLISHER_LABELS:
            raise ValueError(f"Unsupported publisher/provider in journal list: {provider!r}")
        if not journal:
            raise ValueError("Every enabled journal must have a Journal title")

        raw_issns = [val(row, "ISSN", "Print ISSN"), val(row, "eISSN", "Electronic ISSN", "Online ISSN")]
        issns: list[str] = []
        for raw in raw_issns:
            # Also allow multiple ISSNs in one cell separated by comma/semicolon/slash.
            for part in re.split(r"[,;/|\n]", str(raw or "")):
                n = normalize_issn(part)
                if n and n not in issns:
                    issns.append(n)

        aliases = _split_aliases(str(val(row, "Aliases", "Alias", "Alternative Titles")))
        specs.append(
            JournalSpec(
                provider=provider,
                publisher=_PUBLISHER_LABELS[provider],
                journal=journal,
                issns=tuple(issns),
                aliases=aliases,
                category=normalize_space(str(val(row, "Category"))),
                notes=normalize_space(str(val(row, "Notes"))),
            )
        )

    return tuple(specs)


def enabled_journals(provider: str) -> tuple[JournalSpec, ...]:
    p = _provider(provider)
    return tuple(j for j in load_journal_list() if j.provider == p)


def extract_issns(value: str | None) -> set[str]:
    text = str(value or "")
    candidates = re.findall(r"\b\d{4}-?\d{3}[\dXx]\b", text)
    if not candidates:
        n = normalize_issn(text)
        return {n} if len(n) == 8 else set()
    return {normalize_issn(v) for v in candidates if len(normalize_issn(v)) == 8}


def match_journal(provider: str, journal: str | None, issn: str | None, specs: Sequence[JournalSpec] | None = None) -> JournalSpec | None:
    candidates = tuple(specs) if specs is not None else enabled_journals(provider)
    if not candidates:
        return None

    record_issns = extract_issns(issn)
    if record_issns:
        for spec in candidates:
            if record_issns.intersection(spec.issns):
                return spec

    title_key = _title_key(journal)
    if title_key:
        for spec in candidates:
            names = (spec.journal, *spec.aliases)
            if title_key in {_title_key(name) for name in names if name}:
                return spec
    return None


def is_allowed_article(provider: str, journal: str | None, issn: str | None) -> bool:
    return match_journal(provider, journal, issn) is not None


def describe_journals(specs: Iterable[JournalSpec]) -> str:
    specs = tuple(specs)
    return ", ".join(f"{s.journal} ({'/'.join(display_issn(x) for x in s.issns) or 'title match'})" for s in specs)
