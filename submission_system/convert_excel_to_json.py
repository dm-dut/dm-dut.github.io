#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Convert journal submission Excel to JSON.

Recommended usage:
    python convert_excel_to_json.py --excel journal_submission_systems.xlsx --out journal_submission_systems.json

If --excel is omitted, the script will try to use:
    journal_submission_systems.xlsx
in the current folder.

Expected columns, Chinese or English headers are both supported:
    分类/category
    系统类型/system_type
    期刊/平台名称/name_cn/name
    英文名称/name_en
    缩写/abbr
    出版社/平台/platform
    链接/url
    原始标题/original_title
    备注/notes

Only the following fields are required:
    期刊/平台名称 or 英文名称
    链接
"""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("错误：缺少 openpyxl。请先运行：", file=sys.stderr)
    print("    pip install openpyxl", file=sys.stderr)
    sys.exit(1)


COLUMN_ALIASES = {
    "category": ["分类", "category", "Category"],
    "system_type": ["系统类型", "system_type", "System Type", "系统"],
    "name_cn": ["期刊/平台名称", "期刊名", "期刊名称", "名称", "name_cn", "name", "Name"],
    "name_en": ["英文名称", "英文名", "name_en", "English Name"],
    "abbr": ["缩写", "abbr", "Abbr", "Abbreviation"],
    "platform": ["出版社/平台", "平台", "platform", "Publisher/Platform"],
    "url": ["链接", "投稿链接", "url", "URL", "link", "Link"],
    "original_title": ["原始标题", "original_title", "Original Title"],
    "notes": ["备注", "notes", "Notes"],
}

DEFAULTS = {
    "category": "未分类",
    "system_type": "",
    "name_cn": "",
    "name_en": "",
    "abbr": "",
    "platform": "",
    "url": "",
    "original_title": "",
    "notes": "",
}


def cell_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none"}:
        return ""
    return text


def find_column(headers, aliases):
    normalized = {h.strip(): i for i, h in enumerate(headers) if h and h.strip()}
    for alias in aliases:
        if alias in normalized:
            return normalized[alias]
    # Case-insensitive fallback
    lower_map = {h.lower(): i for h, i in normalized.items()}
    for alias in aliases:
        if alias.lower() in lower_map:
            return lower_map[alias.lower()]
    return None


def choose_sheet(workbook):
    if "SubmissionLinks" in workbook.sheetnames:
        return workbook["SubmissionLinks"]
    # Prefer a sheet with a URL/link-like header
    for ws in workbook.worksheets:
        headers = [cell_text(c.value) for c in ws[1]]
        if find_column(headers, COLUMN_ALIASES["url"]) is not None:
            return ws
    return workbook.active


def convert(excel_path: Path, out_path: Path):
    if not excel_path.exists():
        print(f"错误：找不到 Excel 文件：{excel_path}", file=sys.stderr)
        print("请检查文件名，或使用 --excel 指定完整路径。", file=sys.stderr)
        sys.exit(1)

    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as exc:
        print(f"错误：无法读取 Excel 文件：{excel_path}", file=sys.stderr)
        print(f"详细信息：{exc}", file=sys.stderr)
        sys.exit(1)

    ws = choose_sheet(wb)
    headers = [cell_text(c.value) for c in ws[1]]

    idx = {}
    for field, aliases in COLUMN_ALIASES.items():
        idx[field] = find_column(headers, aliases)

    missing_required = []
    if idx["url"] is None:
        missing_required.append("链接/url")
    if idx["name_cn"] is None and idx["name_en"] is None:
        missing_required.append("期刊/平台名称/name 或 英文名称/name_en")

    if missing_required:
        print("错误：Excel 缺少必要列：", "，".join(missing_required), file=sys.stderr)
        print("当前识别到的表头：", " | ".join(headers), file=sys.stderr)
        sys.exit(1)

    links = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        item = dict(DEFAULTS)
        for field in DEFAULTS:
            col = idx.get(field)
            if col is not None and col < len(row):
                item[field] = cell_text(row[col])

        if not item["url"]:
            continue

        # Web page uses name_cn first, then name_en.
        if not item["name_cn"] and item["name_en"]:
            item["name_cn"] = item["name_en"]
        if not item["name_en"] and item["name_cn"]:
            item["name_en"] = item["name_cn"]

        item["id"] = len(links) + 1
        item["updated_date"] = date.today().isoformat()
        links.append(item)

    data = {
        "generated_at": date.today().isoformat(),
        "count": len(links),
        "links": links,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"已生成 {out_path}，共 {len(links)} 条。")
    print(f"读取工作表：{ws.title}")


def main():
    parser = argparse.ArgumentParser(description="Convert journal submission Excel to JSON.")
    parser.add_argument(
        "--excel",
        default="journal_submission_systems.xlsx",
        help="输入 Excel 文件路径，默认 journal_submission_systems.xlsx",
    )
    parser.add_argument(
        "--out",
        default="journal_submission_systems.json",
        help="输出 JSON 文件路径，默认 journal_submission_systems.json",
    )
    args = parser.parse_args()
    convert(Path(args.excel), Path(args.out))


if __name__ == "__main__":
    main()
