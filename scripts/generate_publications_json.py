#!/usr/bin/env python3
"""Generate data/publications.json from the homepage publication Excel database.

Rules used for the current homepage:
- Preserve both English and Chinese fields; the website can switch display language.
- Exclude working papers (Type == 工作 / Working / Working Paper).
- Include published and formally accepted records, including Chinese-language records
  when English title/source fields are available.
- DOI links are generated as https://doi.org/<DOI>; records without DOI keep link empty.
- English corresponding authors are read from Corresponding_Author_en.
- Chinese corresponding authors are read from Corresponding_Author_cn.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("Please install openpyxl first: pip install openpyxl") from exc


TYPE_MAP = {
    "专著": "Monograph",
    "著作": "Monograph",
    "book": "Monograph",
    "monograph": "Monograph",
    "章节": "Book Chapter",
    "book chapter": "Book Chapter",
    "期刊": "Journal Article",
    "journal": "Journal Article",
    "journal article": "Journal Article",
    "会议": "Conference Paper",
    "conference": "Conference Paper",
    "conference paper": "Conference Paper",
}

WORKING_TYPES = {
    "工作",
    "working",
    "working paper",
    "work",
    "working papers",
}


def clean(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).replace("\r", " ").replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    if s.lower() in {"nan", "none"}:
        return ""
    return s


def norm_key(v: Any) -> str:
    return clean(v).strip().lower()


def norm_name(name: str) -> str:
    """Normalize English-style names for display.

    The homepage uses the inverted author-name style consistently for
    English names, e.g., ``Zhang, Zhen``. Chinese names are left unchanged.
    Existing inverted names are preserved; non-inverted English names such as
    ``Zhen Zhang`` are converted to ``Zhang, Zhen`` when possible.
    """
    s = clean(name).replace("*", "")
    s = re.sub(r"\s+", " ", s).strip()

    if not s or re.search(r"[\u4e00-\u9fff]", s):
        return s

    if "," in s:
        family, given = [p.strip() for p in s.split(",", 1)]
        return f"{family}, {given}" if family and given else s

    parts = s.split()
    if len(parts) >= 2:
        family = parts[-1]
        given = " ".join(parts[:-1])
        return f"{family}, {given}"

    return s


def name_key(name: str) -> str:
    """Canonical key for English author matching.

    This key treats ``Zhang, Zhen`` and ``Zhen Zhang`` as the same person.
    """
    s = clean(name).replace("*", "")
    s = re.sub(r"\s+", " ", s).strip().lower()

    if not s:
        return ""

    if "," in s:
        family, given = [p.strip() for p in s.split(",", 1)]
        canonical = family + given
    else:
        parts = s.split()
        canonical = (parts[-1] + "".join(parts[:-1])) if len(parts) >= 2 else s

    return re.sub(r"[^a-z]", "", canonical)


def zh_key(name: str) -> str:
    """Chinese-name key; keeps only CJK characters."""
    return "".join(re.findall(r"[\u4e00-\u9fff]", clean(name).replace("*", "")))


def split_author_cell(authors: str) -> list[str]:
    """Split author lists.

    Author cells should use semicolon or Chinese semicolon between authors.
    English names may contain commas, so commas are not used as separators here.
    """
    return [p.strip() for p in re.split(r";|；", clean(authors)) if p.strip()]


def split_corresponding_en_cell(corresponding: str) -> list[str]:
    """Split English corresponding-author cells.

    English names may be stored as ``Family, Given``, so a simple comma should
    not be treated as a separator.
    """
    s = clean(corresponding)
    if not s:
        return []
    return [
        p.strip()
        for p in re.split(r";|；|\band\b|,\s+and\s+", s, flags=re.I)
        if p.strip()
    ]


def split_corresponding_cn_cell(corresponding: str) -> list[str]:
    """Split Chinese corresponding-author cells."""
    s = clean(corresponding)
    if not s:
        return []
    return [
        p.strip()
        for p in re.split(r";|；|,|，|、|\band\b", s, flags=re.I)
        if p.strip()
    ]


def mark_english_corresponding_authors(
    authors_en: str,
    corresponding_en: str,
) -> list[str]:
    """Return English author list with ``*`` after corresponding authors."""
    authors = split_author_cell(authors_en)
    corresponding_keys = {
        name_key(x)
        for x in split_corresponding_en_cell(corresponding_en)
        if name_key(x)
    }

    output = []

    for author in authors:
        display_name = norm_name(author)
        key = name_key(author)

        if key and key in corresponding_keys and not display_name.endswith("*"):
            display_name += "*"

        output.append(display_name)

    return output


def mark_chinese_corresponding_authors(
    authors_zh: str,
    corresponding_cn: str,
) -> list[str]:
    """Return Chinese author list with ``*`` after corresponding authors."""
    authors = split_author_cell(authors_zh)
    corresponding_keys = {
        zh_key(x)
        for x in split_corresponding_cn_cell(corresponding_cn)
        if zh_key(x)
    }

    output = []

    for author in authors:
        display_name = clean(author).replace("*", "").strip()
        key = zh_key(author)

        if key and key in corresponding_keys and not display_name.endswith("*"):
            display_name += "*"

        output.append(display_name)

    return output


def split_tags(v: Any) -> list[str]:
    s = clean(v)
    if not s:
        return []
    return [x.strip() for x in re.split(r"[,;]", s) if x.strip()]


def doi_link(doi: str) -> tuple[str, str]:
    d = clean(doi)

    if not d:
        return "", ""

    d = re.sub(r"^https?://(dx\.)?doi\.org/", "", d, flags=re.I).strip()
    return d, f"https://doi.org/{d}"


def year_value(v: Any) -> str:
    s = clean(v)

    if not s:
        return ""

    m = re.search(r"\d{4}", s)
    return m.group(0) if m else s


def build_indexes(row: dict[str, Any]) -> list[str]:
    tags = []

    for key in ["SCI", "SSCI", "EI", "CSSCI", "ISTP"]:
        if clean(row.get(key)):
            tags.append(key)

    fms = clean(row.get("FMS"))
    if fms:
        tags.append(f"FMS {fms}" if not fms.upper().startswith("FMS") else fms)

    abs_key = "ABS " if "ABS " in row else "ABS"
    abs_v = clean(row.get(abs_key))
    if abs_v:
        tags.append(f"ABS {abs_v}" if not abs_v.upper().startswith("ABS") else abs_v)

    if clean(row.get("IF")):
        tags.append(f"2025 IF={clean(row.get('IF'))}")

    if clean(row.get("ESI_Highly_Cited")):
        tags.append("ESI Highly Cited Paper")

    if clean(row.get("ESI_Hot")):
        tags.append("ESI Hot Paper")

    seen = set()
    output = []

    for tag in tags:
        if tag not in seen:
            seen.add(tag)
            output.append(tag)

    return output


def convert(input_xlsx: Path, output_json: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(input_xlsx, data_only=True)
    ws = wb.active

    headers = [
        clean(ws.cell(1, c).value)
        for c in range(1, ws.max_column + 1)
    ]

    records = []

    for r in range(2, ws.max_row + 1):
        row = {
            headers[c - 1]: ws.cell(r, c).value
            for c in range(1, ws.max_column + 1)
        }

        raw_type = norm_key(row.get("Type"))

        if not raw_type or raw_type in WORKING_TYPES:
            continue

        pub_type = TYPE_MAP.get(
            raw_type,
            TYPE_MAP.get(clean(row.get("Type")), "Other"),
        )

        title_en = clean(row.get("Title_English"))
        title_zh = clean(row.get("Title_Chinese"))
        title = title_en or title_zh

        venue_en = clean(row.get("Source_English"))
        venue_zh = clean(row.get("Source_Chinese"))
        venue = venue_en or venue_zh

        corresponding_en = clean(row.get("Corresponding_Author_en"))
        corresponding_cn = clean(row.get("Corresponding_Author_cn"))

        authors_en = mark_english_corresponding_authors(
            clean(row.get("Author_English")),
            corresponding_en,
        )

        authors_zh = mark_chinese_corresponding_authors(
            clean(row.get("Author_Chinese")),
            corresponding_cn,
        )

        authors = authors_en or authors_zh

        if not (title or venue or authors):
            continue

        doi, link = doi_link(clean(row.get("DOI")))

        language_raw = clean(row.get("Language"))

        if "中文" in language_raw or language_raw.lower().startswith("zh"):
            language = "zh"
        elif "英文" in language_raw or language_raw.lower().startswith("en"):
            language = "en"
        else:
            language = language_raw

        rec = {
            "type": pub_type,
            "language": language,
            "year": year_value(row.get("Year")),

            "title": title,
            "title_en": title_en,
            "title_zh": title_zh,

            "authors": authors,
            "authors_en": authors_en,
            "authors_zh": authors_zh,

            "corresponding_author_en": corresponding_en,
            "corresponding_author_cn": corresponding_cn,

            "venue": venue,
            "venue_en": venue_en,
            "venue_zh": venue_zh,

            "volume": clean(row.get("Volume")),
            "issue": clean(row.get("Number")),
            "pages": clean(row.get("Page")),

            "doi": doi,
            "link": link,

            "indexes": build_indexes(row),
            "labels": [],

            "note": clean(row.get("Note_English")),
            "note_en": clean(row.get("Note_English")),
            "note_zh": clean(row.get("Note_Chinese")),

            "isbn": clean(row.get("ISBN")),
            "address": clean(row.get("Address")),
            "conference_date": clean(row.get("Conference_Date")),
            "conference_address": clean(row.get("Conference_Address")),

            "source_row": r,
        }

        records.append(rec)

    type_order = {
        "Monograph": 0,
        "Book Chapter": 1,
        "Journal Article": 2,
        "Conference Paper": 3,
        "Other": 9,
    }

    records.sort(
        key=lambda x: (
            -(
                int(x["year"] or 0)
                if str(x["year"]).isdigit()
                else 0
            ),
            type_order.get(x["type"], 9),
            x.get("venue", ""),
            x.get("title", ""),
        )
    )

    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(
        json.dumps(records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return records


if __name__ == "__main__":
    root = Path(__file__).resolve().parents[1]

    input_xlsx = (
        Path(sys.argv[1])
        if len(sys.argv) > 1
        else root / "publication_database.xlsx"
    )

    output_json = (
        Path(sys.argv[2])
        if len(sys.argv) > 2
        else root / "data" / "publications.json"
    )

    records = convert(input_xlsx, output_json)

    print(f"Generated {len(records)} publication records -> {output_json}")