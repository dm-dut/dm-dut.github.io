# -*- coding: utf-8 -*-
"""Convert journal_list.xlsx to journals.json.

Excel columns:
Category | Order | Journal | URL | Extra Links (label|url; ...) | Note

Example for extra links:
CNKI|https://example.com; Wanfang|https://example.com
"""
import argparse
import json
from pathlib import Path
from openpyxl import load_workbook

EXTRA_COL = 'Extra Links (label|url; ...)'

def clean(value):
    if value is None:
        return ''
    return str(value).strip()

def split_extra_links(text):
    links = []
    text = clean(text)
    if not text:
        return links
    for part in [p.strip() for p in text.split(';') if p.strip()]:
        if '|' in part:
            label, url = part.split('|', 1)
        else:
            label, url = 'Link', part
        label, url = clean(label), clean(url)
        if url:
            links.append({'label': label or 'Link', 'url': url})
    return links

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--excel', default='journal_list.xlsx')
    parser.add_argument('--out', default='journals.json')
    args = parser.parse_args()

    wb = load_workbook(args.excel, data_only=True)
    ws = wb['Journals'] if 'Journals' in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit('No data found in Excel.')

    headers = [clean(h) for h in rows[0]]
    hmap = {h.lower(): i for i, h in enumerate(headers)}
    required = ['category', 'journal', 'url']
    missing = [r for r in required if r not in hmap]
    if missing:
        raise SystemExit('Missing columns: ' + ', '.join(missing))

    def get(row, col):
        idx = hmap.get(col.lower())
        return row[idx] if idx is not None and idx < len(row) else ''

    categories, journals = [], []
    per_cat_counter = {}
    for row in rows[1:]:
        if not any(row):
            continue
        category = clean(get(row, 'Category'))
        journal = clean(get(row, 'Journal'))
        url = clean(get(row, 'URL'))
        if not category or not journal:
            continue
        if category not in categories:
            categories.append(category)
        per_cat_counter[category] = per_cat_counter.get(category, 0) + 1
        order_raw = clean(get(row, 'Order'))
        try:
            order = int(float(order_raw)) if order_raw else per_cat_counter[category]
        except Exception:
            order = per_cat_counter[category]
        journals.append({
            'category': category,
            'order': order,
            'journal': journal,
            'url': url,
            'extra_links': split_extra_links(get(row, EXTRA_COL)),
            'note': clean(get(row, 'Note'))
        })

    journals.sort(key=lambda x: (categories.index(x['category']) if x['category'] in categories else 999, x['order']))
    Path(args.out).write_text(json.dumps({'categories': categories, 'journals': journals}, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'Updated {args.out}: {len(journals)} journals, {len(categories)} categories.')

if __name__ == '__main__':
    main()
